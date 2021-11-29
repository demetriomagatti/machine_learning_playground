import pandas as pd
from openpyxl import load_workbook


class Logger():
    '''
    Object to log and save lite results for a specific configuration.

    Arguments:
        excel_filepath: path to the excel file to look for/create/write results in;
        test routine specifics (auto_train, added_features, ...).
    '''
    def __init__(self,*args,**kwargs):
        # Set default parameters
        self.excel_filepath = 'results/df/test_log.xlsx'
        self.auto_train = False
        self.added_features = 0
        self.oversample = False
        self.features = None
        self.train_size = 0.66
        self.stratify = False
        self.robustness_iterations = 100
        self.df_results = pd.DataFrame()
        self.df_name = ''
        self.query = ''
        # Update with provided parameters
        self.__dict__.update(kwargs)
    
    
    def update_log(self):
        '''
        Update log file.
        '''
        line_info = {
            'auto_train': self.auto_train,
            'added_features': self.added_features,
            'oversample': self.oversample,
            'features': self.features,
            'train_size': self.train_size,
            'stratified': self.stratify,
            'robustness_iterations': self.robustness_iterations,
            'results_df': self.df_name
            }
        line_info.update(self.df_results.mean().to_dict())     
        try:
            book = load_workbook(self.excel_filepath)
            writer = pd.ExcelWriter(self.excel_filepath, engine='openpyxl') 
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            for sheetname in writer.sheets:
                pd.DataFrame([line_info]).to_excel(writer,sheet_name=sheetname,startrow=writer.sheets[sheetname].max_row,
                                                   index=True,header=False)
            writer.save()
        except: 
            print('No test log file - creating one')
            pd.DataFrame([line_info]).to_excel(self.excel_filepath)
            
    
    def find_row(self):
        '''
        Look for a specific configuration in the log file.
        '''
        query_str = ''
        for key in self.query.keys():
            query_str = query_str + f'{key} == @self.query["{key}"] & ' 
        log_df = pd.read_excel(self.excel_filepath).drop(['Unnamed: 0'],axis=1)
        return log_df.query(query_str[:-2])